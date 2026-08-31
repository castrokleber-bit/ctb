"""Critério de aceite da Fase 1.

O dicionário só vale se três coisas forem verdade ao mesmo tempo:

1. **Estrutura** — nenhuma conta é ancestral de outra (dupla contagem) e toda base de
   incidência existe. Verificado no carregamento.
2. **Cobertura** — nenhuma conta do ramo de arrecadação fica órfã em nenhum ano da série.
3. **Reconciliação** — aplicado a 2024, o dicionário reproduz o alvo da opção B, que é
   `principal + acessórios` por rubrica nos microdados do `CTB2024.xlsx`.

Este módulo não conserta nada: ele relata e devolve código de saída diferente de zero.
"""

from __future__ import annotations

import collections
import json
import warnings
from pathlib import Path

from pipeline.dominio.dicionario import (
    ErroDicionario, carregar_mapeamentos, carregar_politica_colunas, classificar,
)
from pipeline.fontes.cache import DIR_CACHE_DCA, esferas_por_ente, itens_dca
from pipeline.fontes.diagnostico import ABAS_MICRODADO, _publicado_uniao, br

# Ramos patrimoniais fora do escopo de arrecadação (não são carga tributária, nunca
# estiveram nos quadros publicados) — ver pipeline/dominio/dicionario.py e
# docs/decisoes-pendentes.md §4. Só para relatório de transparência.
RAMOS_FORA_DE_ESCOPO = ("RO1.3.1.", "RO1.3.2.", "RO1.3.3.", "RO1.3.5.", "RO1.3.6.", "RO1.3.9.")

# Rubrica do dicionário → rótulo da linha em `byGOVDetalhado`, quando os dois não têm o
# mesmo nome. Só para relatório; não entra no cálculo.
ROTULO_PUBLICADO = {
    "Contrib. Seg. Serv. Público": "Contrib. Seg. Serv. Público (2)",
    "Outras contribuições sociais": "Outras contribuições sociais (3)",
    "Contribuições Econômicas": "Contribuições Econômicas (4)",
    "Previdência Social": "Previdência (1)",
}


def _casar_linha(publicado: dict[str, float], rubrica: str) -> float | None:
    for chave in (ROTULO_PUBLICADO.get(rubrica, rubrica), rubrica):
        if chave in publicado:
            return publicado[chave]
    alvo = rubrica.casefold()
    for k, v in publicado.items():
        if k.casefold().startswith(alvo[:12]):
            return v
    return None
from pipeline.fontes.http import RAIZ

# Prefixo de natureza de 8 dígitos (planilha) → rubrica do dicionário. Serve só para
# construir o alvo da reconciliação a partir do CTB2024.xlsx; não é usado no cálculo.
PREFIXO_PARA_RUBRICA = {
    "1111": "Imp. sobre Comércio Exterior",
    "1112": "ITR",
    "1113": "IR",
    "1114": "IPI",
    "1115": "IOF",
    "1119": "Outros impostos",
    "1121": "Taxas",
    "1122": "Taxas",
    "1211": "Cofins",
    "1212": "PIS-PASEP",
    "1213": "CSLL",
    "1214": "Previdência Social",
    # Decisão 4 (2026-08-31): royalties têm linha própria. Estas quatro naturezas são o
    # conteúdo integral da aba `Patrimoniais` do CTB2024.xlsx (royalties de petróleo em
    # concessão e partilha, minerais, participação especial). Esperado NÃO fechar ao
    # centavo contra a DCA: a planilha não inclui outorgas/bônus de assinatura e alguns
    # acessórios que o ramo 1.3.4 da DCA traz — ver docs/divergencias.md.
    "1341": "Royalties e Compensações Financeiras",
    "1343": "Royalties e Compensações Financeiras",
    "1344": "Royalties e Compensações Financeiras",
    "1345": "Royalties e Compensações Financeiras",
}

# Rubricas com residual conhecido e documentado contra o alvo — não contam como falha
# dentro desta tolerância mais larga. Cada uma tem a razão do residual em
# docs/divergencias.md ou docs/decisoes-pendentes.md.
TOLERANCIA_AMPLIADA = {
    "Royalties e Compensações Financeiras": 3.0,  # docs/decisoes-pendentes.md §4
    # decisão 2 (2026-08-31): a COSIP do DF saiu do bloco Estados para o bloco
    # Municípios — R$ 0,323 bi que o publicado antigo conta de um lado e o novo conta do
    # outro. Some com um resíduo pré-existente e já documentado em Municípios
    # (contribuições econômicas municipais fora do total antigo — docs/divergencias.md
    # §6): 0,323 + 0,288 ≈ 0,61 no bloco Municípios. Tolerância cobre os dois lados.
    "Contribuições de Melhoria e Econômicas": 0.7,
}


def _alvo_opcao_b() -> dict[str, float]:
    """Alvo por rubrica: principal + acessórios dos microdados de 2024 da planilha."""
    import openpyxl

    caminho = RAIZ / "CTB2024.xlsx"
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    alvo: dict[str, float] = collections.defaultdict(float)
    try:
        for aba in ABAS_MICRODADO:
            for linha in wb[aba].iter_rows(values_only=True):
                cod, valor = linha[0], linha[2]
                if cod is None or valor is None:
                    continue
                cod = str(cod).strip()
                if cod.isdigit() and len(cod) == 8 and cod[:4] in PREFIXO_PARA_RUBRICA:
                    alvo[PREFIXO_PARA_RUBRICA[cod[:4]]] += float(valor)
    finally:
        wb.close()
    return dict(alvo)


def validar(esfera: str, anos: range, tolerancia_bi: float = 0.1) -> int:
    print(f"Validando o dicionário da esfera {esfera} para {anos.start}–{anos.stop - 1}\n")
    falhas = 0

    # 1. estrutura
    try:
        mapas = carregar_mapeamentos(esfera)
        politica = carregar_politica_colunas()
    except ErroDicionario as e:
        print(f"❌ estrutura\n{e}")
        return 1
    print(f"✅ estrutura: {len(mapas)} contas mapeadas, "
          f"{len({m.rubrica for m in mapas})} rubricas, nenhuma dupla contagem")

    # 2. cobertura
    esferas = esferas_por_ente()
    orfas_por_ano: dict[int, dict[str, int]] = {}
    for ano in anos:
        contagem: dict[str, int] = collections.Counter()
        for itens in itens_dca(ano, esfera, esferas).values():
            _, orfas = classificar(itens, esfera, ano, mapas, politica)
            contagem.update(orfas)
        if contagem:
            orfas_por_ano[ano] = dict(contagem)

    if orfas_por_ano:
        falhas += 1
        total = {c for m in orfas_por_ano.values() for c in m}
        print(f"\n❌ cobertura: {len(total)} conta(s) de arrecadação sem rubrica")
        rotulos = _rotulos(anos, esfera, esferas)
        for conta in sorted(total):
            presente = [str(a) for a in anos if conta in orfas_por_ano.get(a, {})]
            print(f"   {conta:<22} anos {','.join(presente):<30} {rotulos.get(conta, '?')[:52]}")
        print("   Toda conta acima precisa entrar no dicionário ou ser declarada fora do")
        print("   escopo de arrecadação. Órfã não vira 'outros' (regra 3 do CLAUDE.md).")
    else:
        print(f"✅ cobertura: nenhuma conta órfã em {anos.stop - anos.start} anos")

    fora_de_escopo = _contas_em_ramos(anos, esfera, esferas, RAMOS_FORA_DE_ESCOPO)
    if fora_de_escopo:
        print(f"\nℹ️  {len(fora_de_escopo)} conta(s) em ramos patrimoniais fora do escopo "
              "de arrecadação (aluguéis, valores mobiliários, delegação de serviços, "
              "cessão de direitos, demais) — decisão 4. Nem classificadas, nem órfãs: "
              "nunca estiveram nos quadros publicados.")

    # 3. continuidade da série
    if len(anos) >= 3:
        falhas += _serie(esfera, anos, mapas, politica, esferas)

    # 4. reconciliação
    if 2024 in anos:
        if esfera == "U":
            # A União publica só o principal, então o alvo vem dos microdados de
            # 8 dígitos somados por prefixo (principal + acessórios = alvo da opção B).
            falhas += _reconciliar(mapas, politica, esferas, tolerancia_bi)
        else:
            # Estados e municípios publicam a receita bruta, que é a mesma coisa que o
            # dicionário calcula. A comparação pode ser direta contra byGOVDetalhado.
            falhas += _reconciliar_bloco(esfera, mapas, politica, esferas, tolerancia_bi)

    print()
    print("Dicionário aprovado." if not falhas else f"{falhas} verificação(ões) falharam.")
    return 1 if falhas else 0


def _publicado_blocos() -> dict[str, dict[str, float]]:
    """Lê os blocos ESTADOS e MUNICÍPIOS da aba `byGOVDetalhado` (R$ bi, publicado)."""
    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(RAIZ / "CTB2024.xlsx", data_only=True, read_only=True)
    blocos: dict[str, dict[str, float]] = {"E": {}, "M": {}}
    atual = None
    try:
        for linha in wb["byGOVDetalhado"].iter_rows(max_col=2, values_only=True):
            rotulo = str(linha[0]).strip() if linha[0] is not None else ""
            chave = rotulo.upper()
            if chave == "ESTADOS":
                atual = "E"
                continue
            if chave in ("MUNICÍPIOS", "MUNICIPIOS"):
                atual = "M"
                continue
            if chave.startswith("UNIÃO"):
                atual = None
                continue
            if atual and rotulo and isinstance(linha[1], (int, float)):
                blocos[atual][rotulo] = float(linha[1])
    finally:
        wb.close()
    return blocos


def _politica_bruta(politica: dict[tuple[str, str], str]) -> dict[tuple[str, str], str]:
    """Variante da política que ignora deduções — só para a checagem estrutural.

    Decisão 6 (2026-08-31) fez estados e municípios passarem a publicar receita líquida,
    igual à União. Isso significa que o valor **calculado de verdade** diverge de
    propósito do `byGOVDetalhado` publicado (que é bruto) — essa divergência é esperada,
    não é erro. Mas a pergunta que a reconciliação precisa continuar respondendo é outra:
    *as contas estão mapeadas certo?* Essa pergunta só faz sentido bruta contra bruta.
    """
    return {chave: ("ignorar" if op == "subtrair" else op) for chave, op in politica.items()}


def _somar_por_bloco(esfera, mapas, politica, esferas, ano) -> dict[tuple[str, str], float]:
    calculado: dict[tuple[str, str], float] = collections.defaultdict(float)
    for itens in itens_dca(ano, esfera, esferas).values():
        valores, _ = classificar(itens, esfera, ano, mapas, politica, por_bloco=True)
        for chave, v in valores.items():
            calculado[chave] += v
    return calculado


def _reconciliar_bloco(esfera, mapas, politica, esferas, tolerancia_bi: float) -> int:
    publicado = _publicado_blocos()
    politica_bruta = _politica_bruta(politica)
    # Estrutural (bruta): é o que decide pass/fail — confere se as contas mapeadas somam
    # certo, igual à publicação de sempre.
    bruto = _somar_por_bloco(esfera, mapas, politica_bruta, esferas, 2024)
    # Real (líquida, decisão 6): é o que o pipeline efetivamente vai publicar.
    liquido = _somar_por_bloco(esfera, mapas, politica, esferas, 2024)

    # O bloco Municípios publicado inclui a parcela do DF (ISS, IPTU e ITBI), que é
    # declarada na esfera `D` e classificada pelo dicionário dos estados. Sem somá-la
    # aqui, a comparação acusaria como divergência algo que é só a regra do DF.
    contribuicao_df: dict[str, float] = collections.defaultdict(float)
    if esfera == "M":
        mapas_e = carregar_mapeamentos("E")
        politica_e_bruta = _politica_bruta(politica)
        for arq_ente, itens in itens_dca(2024, "E", esferas).items():
            if esferas.get(arq_ente) != "D":
                continue
            valores, _ = classificar(itens, "D", 2024, mapas_e, politica_e_bruta, por_bloco=True)
            for (bloco, rubrica), v in valores.items():
                if bloco == "M":
                    bruto[(bloco, rubrica)] += v
                    contribuicao_df[rubrica] += v
            valores_liq, _ = classificar(itens, "D", 2024, mapas_e, politica, por_bloco=True)
            for (bloco, rubrica), v in valores_liq.items():
                if bloco == "M":
                    liquido[(bloco, rubrica)] += v

    B = 1e9
    fora = 0
    # Só o bloco da própria esfera é comparável com o publicado. O que a esfera E manda
    # para o bloco Municípios é a parcela do DF — componente, não total: comparar contra
    # a linha publicada de Municípios acusaria uma diferença que é só a ausência dos
    # 5.569 municípios. Esse resto é relatado, não conferido.
    proprio = "E" if esfera in ("E", "D") else esfera
    for bloco in sorted({b for b, _ in bruto} | {b for b, _ in liquido}):
        if bloco != proprio:
            print(f"\n   Parcela que a esfera {esfera} envia ao bloco "
                  f"{'Municípios' if bloco == 'M' else 'Estados'} (regra do DF), líquida:")
            for rubrica in sorted(r for b, r in liquido if b == bloco):
                print(f'   {rubrica:<40} {br(liquido[(bloco, rubrica)] / B, 3):>12}')
            continue
        alvo = publicado.get(bloco, {})
        rotulo_bloco = "Estados" if bloco == "E" else "Municípios"
        print(f"\n   Reconciliação estrutural 2024 — bloco {rotulo_bloco}, receita BRUTA "
              f"(contra byGOVDetalhado — este é o teste que passa ou falha):")
        if contribuicao_df:
            print("   (inclui a parcela bruta do DF: "
                  + ", ".join(f"{r} {br(v / B, 3)}" for r, v in sorted(contribuicao_df.items()))
                  + ")")
        cab = f'   {"rubrica":<40} {"dicionário":>12} {"publicado":>12} {"diferença":>11}'
        print(cab)
        print("   " + "-" * (len(cab) - 3))
        rubricas = sorted({r for b, r in bruto if b == bloco} | set(alvo))
        for rubrica in rubricas:
            c = bruto.get((bloco, rubrica), 0.0) / B
            p = _casar_linha(alvo, rubrica)
            if p is None:
                print(f'   {rubrica:<40} {br(c, 3):>12} {"—":>12} {"—":>11}')
                continue
            dif = c - p
            tol = TOLERANCIA_AMPLIADA.get(rubrica, tolerancia_bi)
            marca = "" if abs(dif) < tol else " ⚠️"
            if rubrica in TOLERANCIA_AMPLIADA and abs(dif) < tol:
                marca = " (residual conhecido, dentro do esperado)"
            if abs(dif) >= tol:
                fora += 1
            print(f'   {rubrica:<40} {br(c, 3):>12} {br(p, 3):>12} '
                  f'{("+" if dif >= 0 else "−") + br(abs(dif), 3):>11}{marca}')

        print(f"\n   Efeito da decisão 6 (2026-08-31) — receita LÍQUIDA que será de fato "
              f"publicada, bloco {rotulo_bloco} (informativo, não entra no pass/fail):")
        cab2 = f'   {"rubrica":<40} {"bruta":>12} {"líquida":>12} {"efeito"}'
        print(cab2)
        print("   " + "-" * (len(cab2) - 3))
        total_efeito = 0.0
        for rubrica in rubricas:
            b_ = bruto.get((bloco, rubrica), 0.0) / B
            l_ = liquido.get((bloco, rubrica), 0.0) / B
            efeito = l_ - b_
            total_efeito += efeito
            if abs(efeito) < 1e-3 and abs(b_) < 1e-3:
                continue
            print(f'   {rubrica:<40} {br(b_, 3):>12} {br(l_, 3):>12} '
                  f'{("+" if efeito >= 0 else "−") + br(abs(efeito), 3)}')
        print(f'   {"TOTAL":<40} {"":>12} {"":>12} '
              f'{("+" if total_efeito >= 0 else "−") + br(abs(total_efeito), 3)}')

    if fora:
        print(f"\n❌ reconciliação estrutural: {fora} rubrica(s) fora da tolerância de "
              f"R$ {br(tolerancia_bi, 1)} bi")
        _nota_extracao(esfera)
        return 1
    print(f"\n✅ reconciliação estrutural: todas as rubricas dentro de R$ {br(tolerancia_bi, 1)} bi")
    return 0


def _serie(esfera: str, anos: range, mapas, politica, esferas) -> int:
    """Continuidade da série: o teste que pega mapeamento faltando numa virada de plano.

    Se uma rubrica tem valor em 2021 e em 2023 mas zera em 2022, não é a economia que
    mudou — é o dicionário que perdeu a conta quando ela mudou de código. Esse é o modo
    de falha mais provável desta fase, e o mais silencioso: os totais de um ano isolado
    continuam fechando.
    """
    B = 1e9
    por_ano: dict[int, dict[str, float]] = {}
    entes_por_ano: dict[int, int] = {}
    for ano in anos:
        acumulado: dict[str, float] = collections.defaultdict(float)
        cache = itens_dca(ano, esfera, esferas)
        entes_por_ano[ano] = sum(1 for itens in cache.values() if itens)
        for itens in cache.values():
            valores, _ = classificar(itens, esfera, ano, mapas, politica)
            for rubrica, v in valores.items():
                acumulado[rubrica] += v
        por_ano[ano] = dict(acumulado)

    rubricas = sorted({r for a in por_ano.values() for r in a})
    lista = list(anos)
    print("\n   Série por rubrica (R$ bi):")
    print("   " + f'{"rubrica":<40}' + "".join(f"{a:>10}" for a in lista))
    print("   " + "-" * (40 + 10 * len(lista)))
    print("   " + f'{"entes declarantes":<40}'
          + "".join(f"{entes_por_ano[a]:>10,}".replace(",", ".") for a in lista))
    for rubrica in rubricas:
        celulas = "".join(f"{br(por_ano[a].get(rubrica, 0.0) / B, 1):>10}" for a in lista)
        print(f"   {rubrica:<40}{celulas}")

    # buracos: zero cercado por valor dos dois lados
    buracos = []
    for rubrica in rubricas:
        for i in range(1, len(lista) - 1):
            ant, atual, prox = lista[i - 1], lista[i], lista[i + 1]
            va = por_ano[ant].get(rubrica, 0.0)
            vc = por_ano[atual].get(rubrica, 0.0)
            vp = por_ano[prox].get(rubrica, 0.0)
            if abs(vc) < 1e6 and abs(va) > 1e8 and abs(vp) > 1e8:
                buracos.append((rubrica, atual))
    if buracos:
        print(f"\n❌ continuidade: {len(buracos)} buraco(s) na série")
        for rubrica, ano in buracos:
            print(f"   {rubrica} zera em {ano} mas tem valor em {ano - 1} e {ano + 1} — "
                  "conta mudou de código e o dicionário não acompanhou")
        return 1
    print("\n✅ continuidade: nenhuma rubrica zera no meio da série")
    return 0


def _nota_extracao(esfera: str) -> None:
    """A DCA municipal é retificada continuamente: divergir da planilha é esperado."""
    dir_ano = DIR_CACHE_DCA / "2024"
    if not dir_ano.exists():
        return
    datas = [arq.stat().st_mtime for arq in dir_ano.iterdir()]
    if not datas:
        return
    from datetime import date

    extracao = date.fromtimestamp(min(datas))
    print(f"\n   Cache extraído do Siconfi em {extracao.isoformat()}. A DCA de estados e"
          "\n   municípios é retificada pelos entes depois da entrega, então divergir da"
          "\n   planilha não é necessariamente erro de dicionário — a fonte mudou desde a"
          "\n   extração dela. Ver `docs/divergencias.md` §5.")


def _rotulos(anos: range, esfera: str, esferas: dict[str, str]) -> dict[str, str]:
    rot = {}
    for ano in anos:
        for itens in itens_dca(ano, esfera, esferas).values():
            for i in itens:
                rot.setdefault(i["cod_conta"], i["conta"])
    return rot


def _contas_em_ramos(anos: range, esfera: str, esferas: dict[str, str], ramos) -> set[str]:
    achadas = set()
    for ano in anos:
        for itens in itens_dca(ano, esfera, esferas).values():
            for i in itens:
                if i["cod_conta"].startswith(ramos):
                    achadas.add(i["cod_conta"])
    return achadas


def _reconciliar(mapas, politica, esferas, tolerancia_bi: float) -> int:
    alvo = _alvo_opcao_b()
    calculado: dict[str, float] = collections.defaultdict(float)
    for itens in itens_dca(2024, "U", esferas).values():
        valores, _ = classificar(itens, "U", 2024, mapas, politica)
        for rubrica, v in valores.items():
            calculado[rubrica] += v

    B = 1e9
    print("\n   Reconciliação 2024 contra o alvo da opção B (principal + acessórios):")
    cab = f'   {"rubrica":<30} {"dicionário":>12} {"alvo":>12} {"diferença":>11}'
    print(cab)
    print("   " + "-" * (len(cab) - 3))
    fora = 0
    for rubrica in sorted(alvo):
        c, a = calculado.get(rubrica, 0.0) / B, alvo[rubrica] / B
        dif = c - a
        tol = TOLERANCIA_AMPLIADA.get(rubrica, tolerancia_bi)
        marca = " " if abs(dif) < tol else " ⚠️"
        if rubrica in TOLERANCIA_AMPLIADA and abs(dif) < tol:
            marca = " (residual conhecido, dentro do esperado)"
        if abs(dif) >= tol:
            fora += 1
        print(f'   {rubrica:<30} {br(c, 3):>12} {br(a, 3):>12} '
              f'{("+" if dif >= 0 else "−") + br(abs(dif), 3):>11}{marca}')

    # Rubricas sem alvo por prefixo: comparar com a linha publicada em byGOVDetalhado.
    # Aqui a diferença é esperada — a linha publicada usa só o principal — mas serve para
    # flagrar composição divergente, que é problema de dicionário, não de metodologia.
    sem_alvo = sorted(set(calculado) - set(alvo))
    if sem_alvo:
        publicado = dict(_publicado_uniao(RAIZ / "CTB2024.xlsx"))  # valores em R$ bi
        # Decisão 4 (2026-08-31): royalties saíram de "Contribuições Econômicas" e
        # ganharam linha própria. A linha publicada de 2024 ainda os inclui (metodologia
        # antiga) — descontar aqui, senão a comparação acusa um "erro" que é a própria
        # decisão já tomada. `calculado` está em R$ (não em bi); converte antes de somar.
        royalties_calc_bi = calculado.get("Royalties e Compensações Financeiras", 0.0) / B
        alvo_ce = _casar_linha(publicado, "Contribuições Econômicas")
        if alvo_ce is not None:
            for chave in list(publicado):
                if "CONTRIBU" in chave.upper() and "ECON" in chave.upper():
                    publicado[chave] = alvo_ce - royalties_calc_bi
        print(f'\n   Rubricas sem alvo por prefixo — contra a linha publicada:')
        print(f'   ("Contribuições Econômicas" já descontada dos R$ {br(royalties_calc_bi, 3)} bi '
              'de royalties — decisão 4)')
        cab2 = f'   {"rubrica":<30} {"dicionário":>12} {"publicado":>12} {"diferença":>11}'
        print(cab2)
        print("   " + "-" * (len(cab2) - 3))
        for rubrica in sem_alvo:
            c = calculado[rubrica] / B
            p = _casar_linha(publicado, rubrica)
            if p is None:
                print(f'   {rubrica:<30} {br(c, 3):>12} {"—":>12} {"—":>11}')
                continue
            dif = c - p
            marca = "" if abs(dif) < 1.0 else "  ⚠️ composição a investigar"
            print(f'   {rubrica:<30} {br(c, 3):>12} {br(p, 3):>12} '
                  f'{("+" if dif >= 0 else "−") + br(abs(dif), 3):>11}{marca}')

    if fora:
        print(f"\n❌ reconciliação: {fora} rubrica(s) fora da tolerância de "
              f"R$ {br(tolerancia_bi, 1)} bi")
        return 1
    print(f"\n✅ reconciliação: todas as rubricas dentro de R$ {br(tolerancia_bi, 1)} bi")
    return 0

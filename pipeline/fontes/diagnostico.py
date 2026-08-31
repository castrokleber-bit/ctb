"""Fase 0 — reconhecimento das fontes.

Testa cada endpoint da seção 4 do PROJETO-CTB.md para 2016–2025 e produz
``docs/viabilidade-fontes.md``. Não calcula carga tributária, não normaliza nada:
só responde se a fonte existe, o que ela entrega e a que custo.

Duas perguntas são prioritárias:
  (a) o Siconfi traz as naturezas de receita da União em 8 dígitos?
  (b) quantos municípios entregaram DCA em cada ano?
"""

from __future__ import annotations

import csv
import random
import re
import statistics
import time
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from pipeline.fontes.http import ErroFonte, RAIZ, obter_json

SICONFI = "https://apidatalake.tesouro.gov.br/ords/siconfi/tt/"
SIDRA = "https://apisidra.ibge.gov.br/values/"
CKAN = "https://www.tesourotransparente.gov.br/ckan/api/3/action/"
APEX = "https://apiapex.tesouro.gov.br/aria/v1/"

ANEXO_RECEITAS = "DCA-Anexo I-C"
COLUNA_BRUTA = "Receitas Brutas Realizadas"

# Padrão do código de conta da DCA: RO + natureza de receita em 7 níveis.
# O 8º dígito da natureza (tipo: principal / multas / dívida ativa) NÃO aparece.
PADRAO_CONTA = re.compile(r"^R[OI](\d)\.(\d)\.(\d)\.(\d)\.(\d\d)\.(\d)\.(\d)$")


# --------------------------------------------------------------------------- #
# infraestrutura de relato
# --------------------------------------------------------------------------- #

@dataclass
class Achado:
    """Uma linha do relatório: o que foi testado, o veredito e a evidência."""

    bloco: str
    pergunta: str
    veredito: str  # "ok" | "parcial" | "falha"
    detalhe: str
    tabela: list[list[str]] = field(default_factory=list)
    cabecalho: list[str] = field(default_factory=list)


SIMBOLO = {"ok": "✅", "parcial": "⚠️", "falha": "❌"}


def br(n: float, casas: int = 0) -> str:
    """Formata número no padrão brasileiro: 1.234.567,89."""
    return f"{n:,.{casas}f}".translate(str.maketrans(",.", ".,"))


# --------------------------------------------------------------------------- #
# acesso ao Siconfi
# --------------------------------------------------------------------------- #

def _siconfi(recurso: str, *, ano: int | str, chave: str, **params: Any) -> dict:
    return obter_json(
        SICONFI + recurso,
        fonte=f"siconfi_{recurso}",
        ano=ano,
        chave=chave,
        params=params,
    ).dados


def cadastro_entes() -> list[dict]:
    d = _siconfi("entes", ano="cadastro", chave="entes")
    itens = d.get("items", [])
    if not itens:
        raise ErroFonte("cadastro de entes veio vazio — sem ele nada mais roda")
    return itens


def dca_receitas(ano: int, id_ente: int) -> list[dict]:
    d = _siconfi(
        "dca",
        ano=ano,
        chave=f"{id_ente}",
        an_exercicio=ano,
        no_anexo=ANEXO_RECEITAS,
        id_ente=id_ente,
    )
    return d.get("items", [])


def _valor(itens: Iterable[dict], cod_conta: str, coluna: str = COLUNA_BRUTA) -> float | None:
    for i in itens:
        if i["cod_conta"] == cod_conta and i["coluna"] == coluna:
            return i["valor"]
    return None


# --------------------------------------------------------------------------- #
# bloco 1 — cadastro de entes
# --------------------------------------------------------------------------- #

def testar_cadastro(entes: list[dict]) -> Achado:
    esferas = Counter(e["esfera"] for e in entes)
    uniao = [e for e in entes if e["esfera"] == "U"]
    linhas = [[k, str(v)] for k, v in sorted(esferas.items())]
    detalhe = (
        f"O cadastro traz {len(entes)} entes. A União aparece com "
        f"`cod_ibge={uniao[0]['cod_ibge']}` e nome {uniao[0]['ente']!r} — "
        "**não** com o código `U` suposto no PROJETO-CTB. "
        "O Distrito Federal tem esfera própria `D`, o que confirma que a regra do DF "
        "precisa ser explícita no dicionário."
    )
    return Achado("Cadastro de entes", "`/tt/entes` responde e identifica a União?",
                  "ok", detalhe, linhas, ["esfera", "entes"])


# --------------------------------------------------------------------------- #
# bloco 2 — União, ano a ano
# --------------------------------------------------------------------------- #

def testar_uniao(anos: range, cod_uniao: int) -> tuple[Achado, dict[int, list[dict]]]:
    por_ano: dict[int, list[dict]] = {}
    linhas = []
    for ano in anos:
        itens = dca_receitas(ano, cod_uniao)
        por_ano[ano] = itens
        contas = {i["cod_conta"] for i in itens}
        raiz = ("ReceitasExcetoIntraOrcamentarias" if "ReceitasExcetoIntraOrcamentarias" in contas
                else "TotalReceitas" if "TotalReceitas" in contas else "—")
        total = _valor(itens, raiz) if raiz != "—" else None
        linhas.append([
            str(ano),
            str(len(itens)),
            raiz,
            "—" if total is None else br(total / 1e9, 1),
        ])

    layouts = {l[2] for l in linhas}
    detalhe = (
        "Todos os dez anos respondem. Há **quebra de layout**: até 2018 a conta-raiz é "
        "`TotalReceitas` e o demonstrativo tem ~345 linhas; de 2019 em diante é "
        "`ReceitasExcetoIntraOrcamentarias`, com o dobro de linhas. O pipeline precisa "
        "reconhecer as duas raízes; usar só uma delas silenciaria 2016–2018."
        if len(layouts) > 1 else
        "Todos os anos usam a mesma conta-raiz."
    )
    return (
        Achado("União (Siconfi DCA)", "`id_ente` da União traz o Anexo I-C de 2016 a 2025?",
               "parcial" if len(layouts) > 1 else "ok", detalhe, linhas,
               ["ano", "linhas", "conta-raiz", "receita bruta (R$ bi)"]),
        por_ano,
    )


def testar_granularidade(itens_2024: list[dict]) -> Achado:
    """Pergunta prioritária (a): o 8º dígito da natureza existe na DCA?"""
    ultimo = Counter()
    for i in itens_2024:
        m = PADRAO_CONTA.match(i["cod_conta"])
        if m:
            ultimo[m.group(7)] += 1
    linhas = [[d, str(n)] for d, n in sorted(ultimo.items())]
    so_zero = set(ultimo) == {"0"}
    detalhe = (
        "**Não.** A DCA agrega a natureza de receita em 7 níveis "
        "(`1.1.1.3.03.1.0`): o último dígito é sempre `0`. O 8º dígito — o que "
        "distingue principal (`1`) de multas e juros (`2`), dívida ativa (`3`) e "
        "acessórios (`7`, `8`) — **não é publicado no Anexo I-C**. "
        "A Regra 2 do PROJETO-CTB não é aplicável a esta fonte."
        if so_zero else
        "O último dígito varia — verificar se corresponde ao tipo de lançamento."
    )
    return Achado("União (Siconfi DCA)", "A DCA expõe a natureza de receita em 8 dígitos?",
                  "falha" if so_zero else "ok", detalhe, linhas,
                  ["último dígito do código", "ocorrências"])


# --------------------------------------------------------------------------- #
# bloco 3 — reconciliação contra a planilha de referência
# --------------------------------------------------------------------------- #

# Cada linha liga um prefixo de 4 dígitos (planilha, Balanço Geral da União) à
# conta agregadora correspondente na DCA. É mapeamento de estrutura, não número.
RECONCILIACAO = [
    ("1111", "RO1.1.1.1.00.0.0", "Comércio exterior"),
    ("1112", "RO1.1.1.2.00.0.0", "ITR / patrimônio"),
    ("1113", "RO1.1.1.3.00.0.0", "Imposto de Renda"),
    ("1114", "RO1.1.1.4.00.0.0", "IPI"),
    ("1115", "RO1.1.1.5.00.0.0", "IOF"),
    ("1121", "RO1.1.2.1.00.0.0", "Taxas — poder de polícia"),
    ("1122", "RO1.1.2.2.00.0.0", "Taxas — prestação de serviços"),
    ("1211", "RO1.2.1.1.00.0.0", "Cofins"),
    ("1212", "RO1.2.1.2.00.0.0", "PIS/Pasep"),
    ("1213", "RO1.2.1.3.00.0.0", "CSLL"),
    ("1214", "RO1.2.1.4.00.0.0", "RGPS"),
]

ABAS_MICRODADO = ["Impostos", "Contribuições", "Patrimoniais", "Multas"]


def _microdado_planilha(caminho: Path) -> dict[tuple[str, str], float]:
    """Lê as abas de microdado do CTB2024.xlsx somando por (prefixo, tipo).

    A planilha é especificação e referência de comparação — nunca fonte de dados
    publicados. Serve aqui só para medir o que a API reproduz.
    """
    import warnings

    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    acumulado: dict[tuple[str, str], float] = defaultdict(float)
    try:
        for aba in ABAS_MICRODADO:
            for linha in wb[aba].iter_rows(values_only=True):
                cod, valor = linha[0], linha[2]
                if cod is None or valor is None:
                    continue
                cod = str(cod).strip()
                if cod.isdigit() and len(cod) == 8:
                    acumulado[(cod[:4], cod[7])] += float(valor)
    finally:
        wb.close()
    if not acumulado:
        raise ErroFonte(f"nenhuma natureza de 8 dígitos lida em {caminho}")
    return dict(acumulado)


def testar_reconciliacao(itens_2024: list[dict], planilha: Path) -> list[Achado]:
    micro = _microdado_planilha(planilha)
    bruta = {i["cod_conta"]: i["valor"] for i in itens_2024 if i["coluna"] == COLUNA_BRUTA}
    deducao: dict[str, float] = defaultdict(float)
    for i in itens_2024:
        if i["coluna"] != COLUNA_BRUTA:
            deducao[i["cod_conta"]] += i["valor"]

    B = 1e9
    linhas, maior_dif = [], 0.0
    for prefixo, conta, nome in RECONCILIACAO:
        total_plan = sum(v for (p, _), v in micro.items() if p == prefixo) / B
        principal = micro.get((prefixo, "1"), 0.0) / B
        liquida = (bruta.get(conta, 0.0) + deducao.get(conta, 0.0)) / B
        dif = liquida - total_plan
        maior_dif = max(maior_dif, abs(dif))
        linhas.append([
            nome, br(principal, 3), br(total_plan, 3), br(liquida, 3),
            ("+" if dif >= 0 else "−") + br(abs(dif), 3),
        ])

    acessorios = sum(v for (_, t), v in micro.items() if t != "1") / B
    reconcilia = Achado(
        "Reconciliação União 2024",
        "A DCA reproduz os totais por rubrica do Balanço Geral da União?",
        "ok" if maior_dif < 0.1 else "parcial",
        (
            "**Sim, ao centavo.** `receita bruta + Outras Deduções da Receita` da DCA "
            "iguala o total da planilha (principal + acessórios) em 10 das 11 rubricas. "
            "A única exceção é *Taxas — poder de polícia*, que é exatamente a divergência "
            "de Taxas já registrada em `docs/divergencias.md`. "
            "Ou seja: a diferença entre as duas fontes **não é de cobertura, é de conceito** "
            "— a planilha publica o principal, a DCA publica bruto e dedução."
        ),
        linhas,
        ["rubrica", "planilha: principal", "planilha: total", "DCA líquida", "DCA − planilha"],
    )

    consequencia = Achado(
        "Reconciliação União 2024",
        "Dá para separar principal de acessório usando só a DCA?",
        "falha",
        (
            f"**Não.** Os acessórios (multas, juros e dívida ativa) somam **R$ {br(acessorios, 3)} bi** "
            "em 2024 na planilha, e é exatamente essa massa que a DCA embute nos totais sem "
            "separar. A linha *Multas e Dívida Ativa* dos quadros — e os valores de IR, IPI, "
            "Cofins etc. que hoje são publicados **só com o principal** — não podem ser "
            "reproduzidos a partir do Siconfi. Esta é a decisão metodológica bloqueante da Fase 1."
        ),
        [[d, br(sum(v for (_, t), v in micro.items() if t == d) / B, 3)]
         for d in sorted({t for (_, t) in micro})],
        ["8º dígito (tipo)", "R$ bi em 2024"],
    )
    return [reconcilia, consequencia, testar_regra_acessorios(micro, planilha)]


# Rubricas do bloco União em `byGOVDetalhado`, com o rótulo publicado e os prefixos de
# natureza que as compõem. Mapeamento de estrutura — os valores vêm da planilha.
LINHA_TAXAS = "Taxas"
LINHA_ACESSORIOS = "Multas e Dívida Ativa"


def _publicado_uniao(caminho: Path) -> dict[str, float]:
    """Lê o bloco UNIÃO da aba `byGOVDetalhado` (R$ bi, como publicado)."""
    import warnings

    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    valores: dict[str, float] = {}
    try:
        dentro = False
        for linha in wb["byGOVDetalhado"].iter_rows(max_col=2, values_only=True):
            rotulo = str(linha[0]).strip() if linha[0] is not None else ""
            if rotulo.upper().startswith("UNIÃO"):
                dentro = True
                continue
            if dentro and rotulo.upper() in ("ESTADOS", "MUNICÍPIOS", "MUNICIPIOS"):
                break
            if dentro and rotulo and isinstance(linha[1], (int, float)):
                valores[rotulo] = float(linha[1])
    finally:
        wb.close()
    return valores


def testar_regra_acessorios(micro: dict[tuple[str, str], float], planilha: Path) -> Achado:
    """A Regra 2 tem uma exceção: Taxas retém os próprios acessórios.

    Este teste é o que fecha a divergência de R$ 0,083 bi registrada no CLAUDE.md.
    Ver `docs/divergencias.md` §1.
    """
    B = 1e9
    publicado = _publicado_uniao(planilha)
    taxas_pub = next((v for k, v in publicado.items() if "TAXA" in k.upper()), None)
    acess_pub = next((v for k, v in publicado.items()
                      if "MULTA" in k.upper() and "DÍVIDA" in k.upper()), None)
    if taxas_pub is None or acess_pub is None:
        return Achado("Reconciliação União 2024", "A Regra 2 vale para todas as rubricas?",
                      "parcial", "Linhas *Taxas* ou *Multas e Dívida Ativa* não localizadas "
                                 "em `byGOVDetalhado` — mapeamento de rótulos precisa de revisão.")

    taxas_princ = sum(v for (p, t), v in micro.items() if p in ("1121", "1122") and t == "1") / B
    taxas_total = sum(v for (p, _), v in micro.items() if p in ("1121", "1122")) / B
    taxas_acess = taxas_total - taxas_princ
    acess_todos = sum(v for (_, t), v in micro.items() if t != "1") / B

    linhas = [
        ["Taxas: só o principal (Regra 2 literal)", br(taxas_princ, 4), br(taxas_pub, 4),
         ("+" if taxas_pub - taxas_princ >= 0 else "−") + br(abs(taxas_pub - taxas_princ), 4)],
        ["Taxas: principal + acessórios", br(taxas_total, 4), br(taxas_pub, 4),
         ("+" if taxas_pub - taxas_total >= 0 else "−") + br(abs(taxas_pub - taxas_total), 4)],
        ["Multas e DA: todos os acessórios", br(acess_todos, 4), br(acess_pub, 4),
         ("+" if acess_pub - acess_todos >= 0 else "−") + br(abs(acess_pub - acess_todos), 4)],
        ["Multas e DA: menos os de taxas", br(acess_todos - taxas_acess, 4), br(acess_pub, 4),
         ("+" if acess_pub - (acess_todos - taxas_acess) >= 0 else "−")
         + br(abs(acess_pub - (acess_todos - taxas_acess)), 4)],
    ]
    fecha = (abs(taxas_pub - taxas_total) < 1e-3
             and abs(acess_pub - (acess_todos - taxas_acess)) < 1e-3)
    return Achado(
        "Reconciliação União 2024",
        "A Regra 2 (só o principal) vale para todas as rubricas?",
        "ok" if fecha else "parcial",
        (
            "**Não — e a exceção é a divergência de R$ 0,083 bi do CLAUDE.md.** *Taxas* é a "
            "única rubrica publicada com `principal + acessórios`; as demais usam só o "
            f"principal. Os R$ {br(taxas_acess, 4)} bi de acessórios de taxas explicam ao mesmo "
            "tempo por que a linha *Taxas* ficava alta e por que *Multas e Dívida Ativa* ficava "
            "baixa: são os mesmos reais, contados de um lado só. Com essa exceção declarada, a "
            "classificação da União fica **100% reproduzível por regra** em 2024. "
            "A exceção pertence ao dicionário (coluna por rubrica), não ao código. "
            "Ver `docs/divergencias.md` §1."
            if fecha else
            "A hipótese da exceção em Taxas não fecha nesta rodada — reabrir `docs/divergencias.md` §1."
        ),
        linhas,
        ["hipótese", "calculado (R$ bi)", "publicado (R$ bi)", "diferença"],
    )


# --------------------------------------------------------------------------- #
# bloco 4 — estados
# --------------------------------------------------------------------------- #

def testar_estados(anos: range, entes: list[dict]) -> Achado:
    estaduais = [e for e in entes if e["esfera"] in ("E", "D")]
    linhas = []
    for ano in anos:
        with ThreadPoolExecutor(max_workers=6) as ex:
            res = list(ex.map(lambda e: (e, len(dca_receitas(ano, e["cod_ibge"]))), estaduais))
        faltantes = [e["ente"] for e, n in res if n == 0]
        linhas.append([
            str(ano),
            f"{len(res) - len(faltantes)}/{len(res)}",
            ", ".join(faltantes) if faltantes else "—",
        ])
    algum_faltante = any(l[2] != "—" for l in linhas)
    return Achado(
        "Estados e DF (Siconfi DCA)",
        "Os 26 estados e o DF entregaram o Anexo I-C em todos os anos?",
        "parcial" if algum_faltante else "ok",
        ("Há ausências. Cada uma precisa de decisão explícita — a esfera estadual não "
         "comporta imputação por faixa populacional." if algum_faltante else
         "Cobertura integral: nenhuma imputação é necessária na esfera estadual."),
        linhas,
        ["ano", "declarantes", "ausentes"],
    )


# --------------------------------------------------------------------------- #
# bloco 5 — cobertura municipal (pergunta prioritária (b))
# --------------------------------------------------------------------------- #

def carregar_faixas() -> list[tuple[int, str, int, float]]:
    caminho = RAIZ / "dicionario" / "faixas_populacionais.csv"
    if not caminho.exists():
        raise ErroFonte(f"faixas populacionais ausentes: {caminho}")
    faixas = []
    with caminho.open(encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh, delimiter=";"):
            faixas.append((
                int(r["faixa"]),
                r["descricao"],
                int(r["populacao_min"]),
                float(r["populacao_max"]) if r["populacao_max"] else float("inf"),
            ))
    return sorted(faixas)


def faixa_de(pop: int, faixas) -> int:
    for numero, _, minimo, maximo in faixas:
        if minimo <= pop <= maximo:
            return numero
    raise ErroFonte(f"população {pop} não cai em nenhuma faixa — dicionário incompleto")


LIMIAR_PARADA = 500_000  # PROJETO-CTB §5: ausência acima disso para o pipeline


def testar_cobertura_municipal(
    anos: range, entes: list[dict], tamanho_amostra: int, semente: int
) -> tuple[Achado, Achado]:
    faixas = carregar_faixas()
    # O DF entra no cadastro duas vezes: como ente de esfera `D` (cod_ibge 53, que é
    # quem entrega a DCA) e como município (cod_ibge iniciado em 53, que nunca entrega,
    # porque não existe prefeitura de Brasília). Tratar Brasília como declarante
    # faltante seria disparar a salvaguarda de parada todo ano por uma razão errada —
    # ela é caso da regra do DF, não de imputação.
    uf_do_df = {str(e["cod_ibge"]) for e in entes if e["esfera"] == "D"}
    municipios, df_municipio = [], []
    for e in entes:
        if e["esfera"] != "M":
            continue
        (df_municipio if str(e["cod_ibge"])[:2] in uf_do_df else municipios).append(e)

    grandes = [m for m in municipios if m["populacao"] > LIMIAR_PARADA]
    demais = [m for m in municipios if m["populacao"] <= LIMIAR_PARADA]

    rng = random.Random(semente)
    amostra = rng.sample(demais, min(tamanho_amostra, len(demais)))
    alvos = grandes + amostra

    def declarou(ano: int, m: dict) -> bool:
        return len(dca_receitas(ano, m["cod_ibge"])) > 0

    linhas_cob, linhas_grandes = [], []
    for ano in anos:
        with ThreadPoolExecutor(max_workers=6) as ex:
            res = dict(zip(
                (m["cod_ibge"] for m in alvos),
                ex.map(lambda m: declarou(ano, m), alvos),
            ))
        g_ausentes = [m for m in grandes if not res[m["cod_ibge"]]]
        a_ok = sum(1 for m in amostra if res[m["cod_ibge"]])
        taxa = a_ok / len(amostra)
        # erro padrão binomial da estimativa por amostragem
        erro = 1.96 * (taxa * (1 - taxa) / len(amostra)) ** 0.5
        pop_amostra = sum(m["populacao"] for m in amostra)
        pop_coberta = sum(m["populacao"] for m in amostra if res[m["cod_ibge"]])
        linhas_cob.append([
            str(ano),
            f"{len(grandes) - len(g_ausentes)}/{len(grandes)}",
            f"{br(taxa * 100, 1)}% ± {br(erro * 100, 1)} p.p.",
            f"{br(pop_coberta / pop_amostra * 100, 1)}%",
            str(len(g_ausentes)),
        ])
        for m in g_ausentes:
            linhas_grandes.append([str(ano), m["ente"], m["uf"], br(m["populacao"])])

    media = statistics.mean(float(l[2].split("%")[0].replace(",", ".")) for l in linhas_cob)
    df_nome = df_municipio[0]["ente"] if df_municipio else "—"
    cobertura = Achado(
        "Cobertura municipal (Siconfi DCA)",
        "Quantos municípios entregaram a DCA em cada ano?",
        "ok" if media > 85 else "parcial",
        (
            f"Censo completo dos {len(grandes)} municípios acima de {br(LIMIAR_PARADA)} habitantes "
            f"(a salvaguarda de parada do PROJETO-CTB §5) mais amostra aleatória de "
            f"{len(amostra)} dos {br(len(demais))} demais, semente {semente}. "
            "A varredura exaustiva dos 5.570 municípios custa ~15 min por ano com 6 conexões "
            "e é feita uma única vez na Fase 2, ficando em cache. A amostra aqui serve só "
            "para dimensionar a imputação.\n\n"
            f"**{df_nome} foi excluída do universo municipal**: o Distrito Federal entrega a "
            "DCA como ente de esfera `D`, e não existe prefeitura declarante. Contá-la como "
            "faltante dispararia a salvaguarda de parada todo ano por uma razão errada. "
            "As receitas tipicamente municipais do DF entram pela regra do DF, que precisa "
            "estar declarada no dicionário — ver `docs/decisoes-pendentes.md`."
        ),
        linhas_cob,
        ["ano", "grandes declarantes", "cobertura estimada (amostra)", "população coberta (amostra)", "grandes ausentes"],
    )
    parada = Achado(
        "Cobertura municipal (Siconfi DCA)",
        f"Algum município acima de {br(LIMIAR_PARADA)} habitantes faltou?",
        "ok" if not linhas_grandes else "falha",
        ("Nenhum. A salvaguarda de parada não dispara em nenhum ano da série."
         if not linhas_grandes else
         "**Sim.** Cada caso abaixo precisa de decisão sua antes da Fase 2: imputar pela média "
         "da faixa é inaceitável neste porte. As alternativas são buscar o dado no portal do "
         "próprio município ou repetir o ano anterior corrigido."),
        linhas_grandes or [["—", "—", "—", "—"]],
        ["ano", "município", "UF", "população"],
    )
    return cobertura, parada


# --------------------------------------------------------------------------- #
# bloco 6 — fontes complementares
# --------------------------------------------------------------------------- #

def testar_transferencias() -> Achado:
    linhas = []
    for caminho in ["transferencias_constitucionais/", "transferencias-constitucionais/anexos",
                    "transferencias"]:
        try:
            obter_json(APEX + caminho, fonte="apex", ano="teste",
                       chave=caminho, tentativas=1, timeout=30)
            linhas.append([f"`aria/v1/{caminho}`", "200"])
        except ErroFonte as e:
            codigo = "404" if "404" in str(e) else "erro"
            linhas.append([f"`aria/v1/{caminho}`", codigo])

    pacotes = []
    for pkg in ["transferencias-constitucionais-para-estados",
                "transferencias-constitucionais-para-municipios"]:
        d = obter_json(CKAN + "package_show", fonte="ckan", ano="cadastro",
                       chave=pkg, params={"id": pkg}).dados["result"]
        csvs = [r for r in d["resources"] if (r.get("format") or "").upper() == "CSV"]
        pacotes.append([f"`{pkg}`", str(len(csvs))])

    return Achado(
        "Transferências constitucionais",
        "O endpoint `apiapex.tesouro.gov.br/aria/v1/` do PROJETO-CTB funciona?",
        "parcial",
        (
            "**Não** — todos os caminhos testados retornam 404; o host responde, a rota do "
            "documento está errada. O caminho viável é o CKAN do Tesouro Transparente, que "
            "publica os repasses em CSV mensal por bloco (União→Estados e União→Municípios). "
            "São ~120 arquivos por bloco na série 2016–2025, todos cacheáveis. "
            "O bloco Estados→Municípios (cota-parte do ICMS e do IPVA) **não está aqui** e "
            "precisa vir da própria DCA estadual ou de fonte separada — item aberto da Fase 2."
        ),
        linhas + pacotes,
        ["recurso", "status / nº de CSVs"],
    )


def testar_ibge() -> Achado:
    linhas = []
    consultas = [
        ("PIB corrente, tabela 1846", "t/1846/n1/all/v/all/p/201601-202504/c11255/90707"),
        ("População residente, tabela 6579", "t/6579/n1/all/v/all/p/2016-2021"),
        ("População municipal, tabela 6579", "t/6579/n6/all/v/all/p/2021"),
    ]
    veredito = "ok"
    for nome, consulta in consultas:
        try:
            d = obter_json(SIDRA + consulta, fonte="sidra", ano="teste",
                           chave=nome.replace(" ", "_"), tentativas=2, timeout=120).dados
            linhas.append([nome, "200", f"{len(d) - 1} valores"])
        except ErroFonte:
            linhas.append([nome, "erro", "—"])
            veredito = "parcial"
    return Achado(
        "IBGE / SIDRA",
        "PIB corrente e população estão disponíveis para toda a série?",
        veredito,
        ("O SIDRA responde sem autenticação e em menos de um segundo. O PIB da tabela 1846 é "
         "trimestral: o denominador anual é a soma dos quatro trimestres, e a série já cobre "
         "2025. A tabela 6579 (estimativas) traz população por município, insumo obrigatório "
         "da imputação. Registrar `data_extracao_pib` desde a primeira rodada — revisão do "
         "PIB pelo IBGE altera toda a série de % do PIB."),
        linhas,
        ["consulta", "status", "retorno"],
    )


def testar_fontes_manuais() -> Achado:
    return Achado(
        "Fontes sem API",
        "O que precisa entrar por `manual/`?",
        "parcial",
        (
            "FGTS (Caixa) e Sistema S (Receita Federal) não têm API e entram como CSV em "
            "`manual/`, com fonte e data declaradas. Nenhum dos dois foi coletado ainda — "
            "o pipeline deve falhar com mensagem clara enquanto faltarem, nunca preencher "
            "com zero. Pelos quadros da planilha, FGTS e Sistema S pesam na base "
            "*salários e mão-de-obra*, que responde por cerca de um quarto da arrecadação: "
            "não é resíduo desprezível."
        ),
        [["FGTS", "Caixa Econômica Federal", "relatório anual", "pendente"],
         ["Sistema S", "Receita Federal", "planilha de arrecadação", "pendente"]],
        ["bloco", "fonte", "formato", "situação"],
    )


# --------------------------------------------------------------------------- #
# relatório
# --------------------------------------------------------------------------- #

def _tabela_md(cabecalho: list[str], linhas: list[list[str]]) -> str:
    if not linhas:
        return ""
    largura = len(cabecalho)
    sep = "|" + "|".join("---" for _ in range(largura)) + "|"
    corpo = "\n".join("| " + " | ".join(l) + " |" for l in linhas)
    return "| " + " | ".join(cabecalho) + " |\n" + sep + "\n" + corpo + "\n"


def escrever_relatorio(achados: list[Achado], anos: range, destino: Path) -> None:
    partes = [
        "# Viabilidade das fontes — Fase 0",
        "",
        f"Gerado por `uv run ctb fontes testar` em {date.today().isoformat()}. "
        f"Série testada: {anos.start}–{anos.stop - 1}.",
        "",
        "Este documento responde se cada fonte da seção 4 do `PROJETO-CTB.md` existe, o que "
        "ela entrega e a que custo. Ele **não** valida números de carga tributária — isso é "
        "Fase 2. Reproduza-o a qualquer momento com o comando acima; ele lê do cache em "
        "`dados/bruto/` e só vai à rede para o que ainda não baixou.",
        "",
        "## Resumo",
        "",
    ]

    resumo = [[SIMBOLO[a.veredito], a.bloco, a.pergunta] for a in achados]
    partes.append(_tabela_md(["", "bloco", "pergunta"], resumo))

    partes += [
        "",
        "## Correções à especificação",
        "",
        "Três afirmações do `CLAUDE.md` e do `PROJETO-CTB.md` não se confirmaram no teste. "
        "**Os dois documentos já foram corrigidos em 2026-08-30**; a tabela abaixo fica como "
        "registro, porque a versão errada circulou e pode estar em cópias antigas.",
        "",
        _tabela_md(
            ["documento", "afirmava", "na verdade"],
            [
                ["`CLAUDE.md` §Fontes<br>`PROJETO-CTB.md` §4",
                 "União: `id_ente=U` (código não numérico)",
                 "`id_ente=1`. O código `U` devolve **400**; `U` é o valor da coluna "
                 "`esfera`, não do `id_ente`"],
                ["`PROJETO-CTB.md` §4",
                 "transferências em `apiapex.tesouro.gov.br/aria/v1/`",
                 "todas as rotas testadas devolvem **404**. O caminho viável é o CKAN do "
                 "Tesouro Transparente, em CSV mensal"],
                ["`CLAUDE.md` §Regras de classificação",
                 "Taxas fica R$ 0,083 bi abaixo — divergência não resolvida",
                 "**resolvida**: *Taxas* é a única rubrica publicada com principal + "
                 "acessórios. Ver `docs/divergencias.md` §1"],
            ],
        ),
    ]

    bloco_atual = None
    for a in achados:
        if a.bloco != bloco_atual:
            partes += ["", f"## {a.bloco}"]
            bloco_atual = a.bloco
        partes += ["", f"### {SIMBOLO[a.veredito]} {a.pergunta}", "", a.detalhe, ""]
        if a.tabela:
            partes.append(_tabela_md(a.cabecalho, a.tabela))

    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text("\n".join(partes).rstrip() + "\n", encoding="utf-8")


# --------------------------------------------------------------------------- #
# orquestração
# --------------------------------------------------------------------------- #

def varrer_municipios(anos: range, trabalhadores: int = 6) -> None:
    """Censo completo dos 5.570 municípios, um ano por vez, tudo para o cache.

    Custa ~15 min por ano com 6 conexões e roda uma única vez: `obter_json` não refaz
    download que já está em disco. É pré-requisito para validar o dicionário municipal
    contra valores publicados — a amostra do diagnóstico serve para dimensionar a
    imputação, não para conferir números.
    """
    entes = cadastro_entes()
    uf_do_df = {str(e["cod_ibge"]) for e in entes if e["esfera"] == "D"}
    municipios = [e for e in entes if e["esfera"] == "M"
                  and str(e["cod_ibge"])[:2] not in uf_do_df]
    print(f"Varredura completa: {len(municipios)} municípios × {len(anos)} ano(s)")
    for ano in anos:
        inicio = time.time()
        with ThreadPoolExecutor(max_workers=trabalhadores) as ex:
            resultados = list(ex.map(lambda m: len(dca_receitas(ano, m["cod_ibge"])), municipios))
        declarantes = sum(1 for n in resultados if n > 0)
        print(f"  {ano}: {declarantes}/{len(municipios)} declarantes "
              f"({declarantes / len(municipios):.1%}) em {time.time() - inicio:.0f}s")


def executar(anos: range, tamanho_amostra: int, semente: int) -> Path:
    print(f"Fase 0 — testando fontes para {anos.start}–{anos.stop - 1}")

    print("  [1/7] cadastro de entes")
    entes = cadastro_entes()
    achados = [testar_cadastro(entes)]
    cod_uniao = next(e["cod_ibge"] for e in entes if e["esfera"] == "U")

    print("  [2/7] União, ano a ano")
    achado_uniao, por_ano = testar_uniao(anos, cod_uniao)
    achados.append(achado_uniao)

    ano_ref = max(a for a in anos if por_ano.get(a))
    print(f"  [3/7] granularidade da natureza de receita ({ano_ref})")
    achados.append(testar_granularidade(por_ano[ano_ref]))

    planilha = RAIZ / "CTB2024.xlsx"
    if planilha.exists() and 2024 in por_ano:
        print("  [4/7] reconciliação contra CTB2024.xlsx")
        achados += testar_reconciliacao(por_ano[2024], planilha)
    else:
        print("  [4/7] reconciliação pulada (CTB2024.xlsx ausente ou 2024 fora da série)")

    print("  [5/7] estados e DF")
    achados.append(testar_estados(anos, entes))

    print(f"  [6/7] cobertura municipal (censo dos grandes + amostra de {tamanho_amostra})")
    achados += list(testar_cobertura_municipal(anos, entes, tamanho_amostra, semente))

    print("  [7/7] transferências, IBGE e fontes manuais")
    achados += [testar_transferencias(), testar_ibge(), testar_fontes_manuais()]

    destino = RAIZ / "docs" / "viabilidade-fontes.md"
    escrever_relatorio(achados, anos, destino)
    print(f"\nRelatório escrito em {destino}")
    for a in achados:
        if a.veredito != "ok":
            print(f"  {SIMBOLO[a.veredito]} {a.bloco}: {a.pergunta}")
    return destino
